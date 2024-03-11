import pyautogui as pygui
from selenium import webdriver
import time
import undetected_chromedriver as uc
from undetected_chromedriver import *
from selenium.webdriver.support.ui import WebDriverWait
import time
import pyautogui
from openpyxl import load_workbook
import pandas as pd

i = 0

driver = uc.Chrome() #define o navegador (undetected chrome)

cpf_list = []
nome_list = []
tipo_exame_list = []
data_exame_list = []
tipo_exame_list_string = ""

values_only = True


#workbook = load_workbook(r'C:\Users\ADMINISTRATOR\Desktop\excelRobo\robo.xlsx')
#worksheet = workbook.active
#rows = worksheet.iter_rows(min_row=1, values_only=True)

worksheet = pd.read_excel(r'C:\Users\ADMINISTRATOR\Desktop\excelRobo\robo.xlsx').to_dict(orient='records')



def fazer_login_um():


    print('iniciando automação!')
    driver.get("https://rh.dtm.com.br/")  # define o url para pesquisa
    driver.maximize_window()  # maximiza a tela

    usuario = 'FORT-46'  # acesso de usuario

    print('inserindo usuario')
    login = driver.find_element('xpath', '//*[@id="Editbox1"]')
    login.click()
    login.send_keys(usuario)
    time.sleep(0.3)

    senha = '3Q7kViZyqt'  # senha de usuario

    print('inserindo senha')
    login_senha = driver.find_element('xpath', '//*[@id="Editbox2"]')
    login_senha.click()
    login_senha.send_keys(senha)
    time.sleep(0.3)

    print('confirmando login')
    clicar_login = driver.find_element('xpath', '//*[@id="buttonLogOn"]')
    clicar_login.click()
    time.sleep(15)

def fazer_login_um_sem_page():

    usuario = 'FORT-46'  # acesso de usuario

    login = driver.find_element('xpath', '//*[@id="Editbox1"]')
    login.click()
    login.send_keys(usuario)
    time.sleep(0.3)

    senha = '3Q7kViZyqt'  # senha de usuario

    login_senha = driver.find_element('xpath', '//*[@id="Editbox2"]')
    login_senha.click()
    login_senha.send_keys(senha)
    time.sleep(0.3)

    clicar_login = driver.find_element('xpath', '//*[@id="buttonLogOn"]')
    clicar_login.click()
    time.sleep(15)

def fazer_login_dois():
    pyautogui.moveTo(734, 419, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print("apagando caso tenha algo escrito")
    pyautogui.typewrite(['backspace'] * 8)
    time.sleep(0.3)
    pyautogui.typewrite("PROSAUDE")
    time.sleep(0.3)  # pausa aqui

    pyautogui.moveTo(729, 440, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print("apagando caso tenha algo escrito")
    pyautogui.typewrite(['backspace'] * 10)
    time.sleep(0.3)
    pyautogui.typewrite("Pro@@2023")
    time.sleep(0.3)
    pyautogui.moveTo(779, 531, duration=2)

    pyautogui.click()
    time.sleep(1)

def exames_medicos():

    print('fechando anuncios')
    time.sleep(4)
    pyautogui.typewrite(['enter'] * 6)
    time.sleep(1)

    print('indo até MO')
    pyautogui.moveTo(222, 150, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('clicando em exames medicos')
    pyautogui.moveTo(1006, 198, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('especifica exames medicos')
    pyautogui.moveTo(311, 239, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('abre exames medicos')
    pyautogui.moveTo(301, 283, duration=2)
    pyautogui.click()
    time.sleep(0.3)

def pesquisa_nome(nome_list):

    pyautogui.moveTo(339, 238, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('abrindo aba para nome')
    pyautogui.moveTo(805, 363, duration=2)
    pyautogui.click()
    print('deixando vazio o espaço para inserção de nome')
    pyautogui.typewrite(['backspace'] * 8)
    time.sleep(0.3)

    pyautogui.typewrite(nome_list, interval=0.7)  # Aqui vai a procedure de nome
    pyautogui.typewrite(['enter'] * 2)
    time.sleep(0.3)

def pesquisa_cpf():

    cpf = cpf_list
    pyautogui.moveTo(639,240, duration=2)
    pyautogui.click()
    pyautogui.click()
    print('inserindo CPF de funcionario')
    pyautogui.typewrite("38426423850",  interval=0.2)
    pyautogui.typewrite(['enter'] * 1)
    time.sleep(0.3)

def sair():

    print('saindo do site')
    pyautogui.typewrite(['esc'] * 7)
    pyautogui.moveTo(1372, 113, duration=2)
    pyautogui.click()
    time.sleep(0.2)
    pyautogui.press('tab')
    time.sleep(0.2)
    pyautogui.press('enter')

x, y = 719, 349

def verificar_cor(x, y):

    cor = pyautogui.pixel(x, y)
    if cor == (132, 130, 255):
        print("login não realizado, realizando login para concluir procedimentos")
        fazer_login_dois()

    else:
        print("Login já realizado, prosseguindo com os procedimentos")
        pyautogui.typewrite(['esc'] * 3)

def verificarCorSecundaria():
    x, y = 133, 600

    print('verificando cor de fundo')
    for i in range(50):
        lugardacor = pyautogui.pixel(x, y)
        if lugardacor == (231,239,247):
            time.sleep(1.5)
            break
        else:
            continue

def incluir():

    print('abrindo incluir')
    pyautogui.moveTo(305, 711, duration=2)
    pyautogui.click()
    time.sleep(2)

def nomeMedico():


    print('indo até nome do medico responsavel')
    pyautogui.moveTo(685, 347, duration=2)
    pyautogui.click()
    time.sleep(1)

    print('localizando barra de pesquisa')
    pyautogui.moveTo(403, 413, duration=2)
    pyautogui.click()
    time.sleep(1)

    print('inserindo nome do medico (FRANCISCO CANHETTI')
    pyautogui.typewrite("FRANCISCO CANHETTI")

    pyautogui.moveTo(511, 564, duration=2)
    pyautogui.click()
    time.sleep(1)

    print('confirmando e fechando')
    pyautogui.moveTo(716, 337, duration=2)
    pyautogui.click()
    pyautogui.click()
    time.sleep(2)

def pesquisaLaboratorio():

    print('indo para laboratorio')
    pyautogui.moveTo(482, 369, duration=2)
    pyautogui.click()
    time.sleep(1)

    pyautogui.moveTo(412, 442, duration=2)
    pyautogui.click()
    time.sleep(1)

    print('pesquisando nome da empresa (PRO OCUPACIONAL)')
    pyautogui.typewrite("PRO OCUPACIONAL")

    pyautogui.moveTo(384, 568, duration=2)
    pyautogui.click()
    time.sleep(1)

    print('confirma e fecha')
    pyautogui.moveTo(791, 336, duration=2)
    pyautogui.click()
    pyautogui.click()
    time.sleep(2)

def dataExameigual(data_exame):

    print("indo para data")
    pyautogui.moveTo(453, 217, duration=2)
    pyautogui.click()
    pyautogui.click()
    time.sleep(0.3)

    print('copiando data')
    pyautogui.typewrite(str(data_exame), interval=0.5)

    time.sleep(0.3)

    pyautogui.moveTo(658, 217, duration=2)
    pyautogui.click()
    pyautogui.click()
    time.sleep(0.3)

    print("colando data")
    pyautogui.typewrite(str(data_exame), interval=0.5)

    time.sleep(0.3)

    pyautogui.moveTo(1059, 396, duration=2)
    pyautogui.click()
    pyautogui.click()
    time.sleep(0.3)

    data = str(data_exame)  # Converta a data para string
    ultimo_digito = int(data[-1])  # Obtenha o último dígito e converta para inteiro
    novo_ultimo_digito = ultimo_digito + 1  # Adicione 1 ao último dígito
    nova_data = data[:-1] + str(
    novo_ultimo_digito)  # Concatene a parte da data sem o último dígito com o novo último dígito
    print("Nova Data do Exame:", nova_data)
#
    pyautogui.typewrite(nova_data, interval=0.5)

    time.sleep(0.3)

def ordemExame():

    print('indo até ordem de exame')
    pyautogui.moveTo(503, 397, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('selecionando sequencial')
    pyautogui.moveTo(460, 439, duration=2)
    pyautogui.click()
    time.sleep(0.3)

def statusExame():

    print('indo até status')
    pyautogui.moveTo(934, 218, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('selecionando finalizado')
    pyautogui.moveTo(931, 246, duration=2)
    pyautogui.click()
    time.sleep(0.3)

def medicoExame():

    print('indo até medico ')
    pyautogui.moveTo(410,324, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('campo de pesquisa')
    pyautogui.moveTo(404,442, duration=2)
    pyautogui.click()
    time.sleep(0.3)

    print('inserindo nome do medico responsavel (FRANCISCO CANHETTI')
    pyautogui.typewrite("FRANCISCO CANHETTI")
    pyautogui.moveTo(389,567, duration=2)
    pyautogui.click()
    time.sleep(1)

    print('confirmando e fechando')
    pyautogui.moveTo(716, 337, duration=2)
    pyautogui.click()
    pyautogui.click()
    time.sleep(2)

def verificaSeTelaAparece():
    x, y =596, 242

    print('verificando cor de fundo')
    for i in range(50):
        lugardacor = pyautogui.pixel(x, y)
        if lugardacor == (0, 121, 214):
            pyautogui.moveTo(1096,112, duration=2)
            pyautogui.click()
            pyautogui.moveTo(693,544, duration=2)
            pyautogui.click()
            pyautogui.typewrite(['esc'] * 4)
            break
        else:
            continue

def fecharPrograma():
    print('fechando programa')
    pyautogui.moveTo(369, 796, duration=2)
    pyautogui.click()
    time.sleep(0.4)
    pyautogui.moveTo(909, 428, duration=2)
    pyautogui.click()
    time.sleep(0.4)
    pyautogui.moveTo(851, 550, duration=2)
    pyautogui.click()
    time.sleep(0.4)
    pyautogui.press(['esc'] * 3)
    time.sleep(1)


def tipoExame(tipo_exame):

    pyautogui.moveTo(718,182, duration=2)
    pyautogui.click()
    time.sleep(1)

    if tipo_exame == "Admissional":
        print("exame admissional realizando o cadastro")
        pyautogui.moveTo(648,201, duration=2)
        pyautogui.click()
        time.sleep(0.3)
    elif tipo_exame == "Demissional":
        print('exame demissional, realizando cadastro')
        pyautogui.moveTo(649,214, duration=2)
        pyautogui.click()
        time.sleep(0.3)
    elif tipo_exame == "Periódico":
        print('exame periodico, realizando cadastro')
        pyautogui.moveTo(642,227, duration=2)
        pyautogui.click()
        time.sleep(0.3)
    else:
        print('nada encontrado')

def processarDados(nome:str, tipo_exame:str, data_exame:str):

    pesquisa_nome(nome)
    time.sleep(1)
    incluir()
    time.sleep(1)
    tipoExame(tipo_exame)
    time.sleep(1)
    dataExameigual(data_exame)
    time.sleep(1)
    ordemExame()
    time.sleep(1)
    statusExame()
    time.sleep(1)
    nomeMedico()
    time.sleep(0.5)
    medicoExame()
    time.sleep(0.5)
    pesquisaLaboratorio()
    time.sleep(1)
    fecharPrograma()
    time.sleep(1)

def buscaExcel():

    next_index = 0
    def formatarData(value):
        return value if len(value) == 8 else '0' + str(value)


    for row in worksheet:
        if next_index == 0:
            row2 = list(row.keys())
            cpf = row2[0]
            nome = row2[1]
            tipo_exame = row2[2]
            data_exame = formatarData(str(row2[3]))
            processarDados(nome, tipo_exame, data_exame)
            next_index += 1
            print("\n" * 1)
            print('*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*')
            print("CPF1:", cpf)
            print("Nome1:", nome)
            print("Tipo de Exame1:", tipo_exame)
            print("Data do Exame1:", data_exame)
            print('*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*_*')
            print("\n" * 1)


        row = list(row.items())
        cpf = row[0][-1]
        nome = row[1][-1]
        tipo_exame = row[2][-1]
        data_exame = row[3][-1]
        processarDados(nome, tipo_exame, data_exame)
        print("\n" * 1)
        print("CPF:", cpf)
        print("Nome:", nome)
        print("Tipo de Exame:", tipo_exame)
        print("Data do Exame:", data_exame)
        print("\n" * 1)


    #$for index, row in enumerate(worksheet.iter_rows(min_row=1, values_only=True), start=1):
    #$    cpf = row[0]
    #$    nome = row[1]
    #$    tipo_exame = row[2]
    #$    data_exame = row[3]






fazer_login_um()
time.sleep(1)
verificarCorSecundaria()
time.sleep(1)
verificar_cor(x, y)
time.sleep(1)
verificarCorSecundaria()
time.sleep(1)
verificaSeTelaAparece()
time.sleep(1)
exames_medicos()
time.sleep(1)
buscaExcel()










